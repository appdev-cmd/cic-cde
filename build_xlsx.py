# -*- coding: utf-8 -*-
"""Mo hinh tai chinh CDE CIC - Cloud-First, tien do giua 2027. Khop 'Tong quan KH tai chinh 5 nam'."""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.comments import Comment

YEARS = ['2026', '2027', '2028', '2029', '2030']
YCOL = ['D', 'E', 'F', 'G', 'H']
NF_TY = '#,##0.0;(#,##0.0);"-"'
NF_INT = '#,##0;(#,##0);"-"'
NF_PCT = '0.0%'
NF_X = '0.00'
NF_TR = '#,##0.00;(#,##0.00);"-"'

BLUE = '0000FF'; GREEN = '008000'; BLACK = '000000'; WHITE = 'FFFFFF'
F = 'Arial'
f_input = Font(name=F, size=10, color=BLUE)
f_link = Font(name=F, size=10, color=GREEN)
f_calc = Font(name=F, size=10, color=BLACK)
f_bold = Font(name=F, size=10, bold=True)
f_boldlink = Font(name=F, size=10, bold=True, color=GREEN)
f_sec = Font(name=F, size=10, bold=True, color=WHITE)
f_sub = Font(name=F, size=9, italic=True, color='808080')
f_hdr = Font(name=F, size=10, bold=True, color=WHITE)
f_title = Font(name=F, size=14, bold=True, color='1F4E78')

fill_survey = PatternFill('solid', fgColor='FFFF00')
fill_input = PatternFill('solid', fgColor='FFF2CC')
fill_sec = PatternFill('solid', fgColor='1F4E78')
fill_hdr = PatternFill('solid', fgColor='4472C4')
fill_tot = PatternFill('solid', fgColor='D9E1F2')
fill_kpi = PatternFill('solid', fgColor='E2EFDA')
nofill = PatternFill(fill_type=None)
thin = Side(style='thin', color='BFBFBF')
border = Border(left=thin, right=thin, top=thin, bottom=thin)
ctr = Alignment(horizontal='center', vertical='center')
left = Alignment(horizontal='left', vertical='center')
rgt = Alignment(horizontal='right', vertical='center')
wrap = Alignment(horizontal='left', vertical='top', wrap_text=True)

wb = Workbook()


def ws_setup(ws, title, subtitle):
    ws.sheet_view.showGridLines = False
    ws['A1'] = title; ws['A1'].font = f_title
    ws['A2'] = subtitle; ws['A2'].font = f_sub
    ws.column_dimensions['A'].width = 11
    ws.column_dimensions['B'].width = 46
    ws.column_dimensions['C'].width = 12
    for c in YCOL:
        ws.column_dimensions[c].width = 12
    ws.column_dimensions['I'].width = 34


def hdr_row(ws, r, extra='Ghi chú'):
    ws[f'A{r}'] = 'Mã'; ws[f'B{r}'] = 'Khoản mục'; ws[f'C{r}'] = 'Đơn vị'
    for i, c in enumerate(YCOL):
        ws[f'{c}{r}'] = YEARS[i]
    ws[f'I{r}'] = extra
    for col in ['A', 'B', 'C'] + YCOL + ['I']:
        cell = ws[f'{col}{r}']; cell.font = f_hdr; cell.fill = fill_hdr
        cell.alignment = ctr; cell.border = border


def section(ws, r, text, span=('A', 'I')):
    ws[f'A{r}'] = text
    cols = [chr(c) for c in range(ord(span[0]), ord(span[1]) + 1)]
    for col in cols:
        ws[f'{col}{r}'].fill = fill_sec
    ws[f'A{r}'].font = f_sec


def setrow(ws, r, code, label, unit, vals, *, kind='input', nf=NF_TY,
           survey=False, note='', total=False, comment=None):
    ws[f'A{r}'] = code; ws[f'A{r}'].font = f_bold; ws[f'A{r}'].alignment = ctr
    ws[f'B{r}'] = label; ws[f'B{r}'].font = (f_bold if total else f_calc); ws[f'B{r}'].alignment = left
    ws[f'C{r}'] = unit; ws[f'C{r}'].font = f_calc; ws[f'C{r}'].alignment = ctr
    fnt = {'input': f_input, 'link': f_link, 'calc': f_calc}[kind]
    if total:
        fnt = Font(name=F, size=10, bold=True, color=(GREEN if kind == 'link' else BLACK))
    if vals is not None:
        for i, c in enumerate(YCOL):
            cell = ws[f'{c}{r}']; cell.value = vals[i]; cell.font = fnt
            cell.number_format = nf; cell.alignment = rgt; cell.border = border
            if kind == 'input':
                cell.fill = fill_survey if survey else fill_input
            elif total:
                cell.fill = fill_tot
    if note:
        ws[f'I{r}'] = note; ws[f'I{r}'].font = f_sub; ws[f'I{r}'].alignment = left
    if comment:
        ws[f'D{r}'].comment = Comment(comment, "Model")
    return r


def no_border_table(ws, r0, r1, c0='A', c1='I'):
    pass  # not needed


# =========================================================================
# 1. HUONG_DAN
# =========================================================================
ws = wb.active; ws.title = 'Huong_Dan'
ws_setup(ws, 'MÔ HÌNH TÀI CHÍNH CDE CIC — HƯỚNG DẪN', 'Cloud-First | Tiến độ 9/2026 – giữa 2027 | Doanh thu từ 2027')
ws.column_dimensions['B'].width = 72
r = 4
notes = [
    ('A. QUY ƯỚC MÀU', 'sec'),
    ('Vàng đậm + chữ xanh', 'Đầu vào CẦN khảo sát lại (có mã)'),
    ('Vàng nhạt + chữ xanh', 'Đầu vào có thể chỉnh (giả định)'),
    ('Chữ xanh lá', 'Công thức liên kết sheet khác'),
    ('Chữ đen', 'Công thức tính trong sheet'),
    ('', ''),
    ('B. NGUYÊN TẮC', 'sec'),
    ('1', 'Chỉ sửa ô MÀU VÀNG ở Gia_Dinh và các sheet chi tiết (Nhan_Su, Thiet_Bi, CP_Khac, Cloud).'),
    ('2', 'Mọi sheet khác là công thức — tự cập nhật. Mở Excel tự tính (nếu trống bấm F9).'),
    ('3', 'Mô hình Cloud-First: KHÔNG mua server (thuê cloud → OPEX). On-Premise: khách tự lo hạ tầng.'),
    ('4', 'NPV dùng quy ước năm 2026 = năm gốc (t=0).'),
    ('', ''),
    ('C. DANH MỤC SHEET', 'sec'),
    ('Thuat_Ngu', 'Chú thích thuật ngữ tiếng Anh & viết tắt'),
    ('Gia_Dinh', 'Toàn bộ giả định đầu vào (nơi điền chính)'),
    ('Nhan_Su', 'Nhân sự bottom-up: đội phát triển (10) + đội vận hành (12/20/22)'),
    ('Thiet_Bi', 'Trang thiết bị văn phòng → CAP-02 (Cloud-First, không mua server)'),
    ('CP_Khac', 'Bản quyền/API, Marketing, PM/Pháp lý → CAP-03/04/05'),
    ('Cloud', 'Chi phí thuê Cloud trong nước → OPX-02'),
    ('Doanh_Thu', 'Doanh thu SaaS + On-Premise (PMU, Sở XD, Doanh nghiệp)'),
    ('Dau_Tu', 'Tổng hợp CAPEX 25,1 tỷ (CIC 70% / NATIF 30%)'),
    ('OPEX', 'Tổng hợp chi phí vận hành'),
    ('P_and_L', 'Báo cáo kết quả & dòng tiền'),
    ('Dong_Tien', 'Dòng tiền dự án & CIC'),
    ('NPV_IRR', 'NPV, IRR, hoàn vốn'),
    ('Hoa_Von', 'Điểm hòa vốn'),
    ('Do_Nhay', 'Phân tích độ nhạy 3 kịch bản'),
    ('Dashboard', 'Bảng tổng hợp chỉ số'),
]
for code, txt in notes:
    if txt == 'sec':
        ws[f'A{r}'] = code; ws[f'A{r}'].font = f_bold; ws[f'A{r}'].fill = fill_tot
        ws[f'B{r}'].fill = fill_tot
    else:
        ws[f'A{r}'] = code; ws[f'B{r}'] = txt
        ws[f'A{r}'].font = f_bold; ws[f'B{r}'].font = f_calc
        if code == 'Vàng đậm + chữ xanh':
            ws[f'A{r}'].fill = fill_survey; ws[f'A{r}'].font = f_input
        elif code == 'Vàng nhạt + chữ xanh':
            ws[f'A{r}'].fill = fill_input; ws[f'A{r}'].font = f_input
    ws[f'A{r}'].alignment = left; ws[f'B{r}'].alignment = left
    r += 1

# =========================================================================
# 2. GIA_DINH
# =========================================================================
gi = {}
ws = wb.create_sheet('Gia_Dinh')
ws_setup(ws, 'GIẢ ĐỊNH ĐẦU VÀO (INPUTS)', 'Chỉ sửa ô vàng. Mỗi dòng có MÃ để tham chiếu khi khảo sát.')
hdr_row(ws, 3)
r = 4


def gset(code, label, unit, vals, **kw):
    global r
    setrow(ws, r, code, label, unit, vals, **kw)
    gi[code] = r; r += 1


def G(code, col):
    return f'Gia_Dinh!{col}{gi[code]}'


section(ws, r, 'A. THAM SỐ TÀI CHÍNH CHUNG'); r += 1
gset('FIN-01', 'Giá vốn dịch vụ (COGS) / doanh thu', '%', [0.40] * 5, nf=NF_PCT, note='⚠️ Triển khai local, băng thông, API, đào tạo')
gset('FIN-02', 'Biên lợi nhuận gộp', '%', [f'=1-D{gi["FIN-01"]}'] * 5, kind='calc', nf=NF_PCT)
gset('FIN-03', 'WACC (tỷ lệ chiết khấu)', '%', [0.12] * 5, nf=NF_PCT, note='✅ Benchmark tech VN')
gset('FIN-04', 'Thuế TNDN', '%', [0.10] * 5, nf=NF_PCT, note='✅ Ưu đãi DN công nghệ')
gset('FIN-05', 'Tỷ lệ vốn CIC (đối ứng)', '%', [0.70] * 5, nf=NF_PCT, note='✅ Cam kết 70/30')
gset('FIN-06', 'Tỷ lệ vốn NATIF (tài trợ)', '%', [f'=1-D{gi["FIN-05"]}'] * 5, kind='calc', nf=NF_PCT)
gset('FIN-07', 'Hệ số tải nhân sự vận hành (overhead)', 'x', [0, 1.62, 1.49, 1.46, 1.46], nf=NF_X,
     note='BHXH + lương 13 + thưởng + tuyển dụng/phúc lợi')

r += 1
section(ws, r, 'B. DOANH THU SaaS'); r += 1
gset('SAAS-01', 'Người dùng active cuối kỳ', 'user', [0, 1000, 4000, 9000, 15000], nf=NF_INT,
     survey=True, note='❌ Sản phẩm ra mắt 2027', comment='2026=0 (đang xây dựng). Thương mại từ giữa 2027.')
gset('SAAS-02', 'ARPU (giá TB/user/tháng)', 'tr', [0, 0.5, 0.5, 0.52, 0.55], nf=NF_TR, survey=True, note='❌ ~6 tr/user/năm')
gset('SAAS-03', 'Số tháng thương mại trong năm', 'tháng', [0, 6, 12, 12, 12], nf=NF_INT, note='2027 từ giữa năm')

r += 1
section(ws, r, 'C. ON-PREMISE — PMU (Ban QLDA)'); r += 1
gset('OPP-01', 'Hợp đồng PMU mới trong năm', 'HĐ', [0, 2, 6, 10, 12], nf=NF_INT, survey=True, note='❌ Lũy kế 30 PMU tới 2030')
gset('OPP-02', 'Giá HĐ PMU (tăng dần)', 'tỷ', [0, 2.0, 2.3, 2.6, 3.0], survey=True, note='❌ 2,0 → 3,0 tỷ')

r += 1
section(ws, r, 'D. ON-PREMISE — SỞ XÂY DỰNG'); r += 1
gset('OSX-01', 'Hợp đồng Sở XD mới trong năm', 'HĐ', [0, 1, 3, 5, 6], nf=NF_INT, survey=True, note='❌ Lũy kế 15 Sở tới 2030')
gset('OSX-02', 'Giá HĐ Sở XD (tăng dần)', 'tỷ', [0, 1.8, 2.0, 2.2, 2.5], survey=True, note='❌ 1,8 → 2,5 tỷ')

r += 1
section(ws, r, 'E. ON-PREMISE — DOANH NGHIỆP LỚN'); r += 1
gset('OPV-01', 'Hợp đồng doanh nghiệp mới', 'HĐ', [0, 1, 3, 4, 5], nf=NF_INT, survey=True, note='⚠️ Lũy kế 13 DN tới 2030')
gset('OPV-02', 'Giá HĐ doanh nghiệp (tăng dần)', 'tỷ', [0, 3.0, 3.4, 3.8, 4.2], survey=True, note='⚠️ 3,0 → 4,2 tỷ')
gset('AMC-01', 'Phí bảo trì hàng năm (% giá HĐ lũy kế)', '%', [0.15] * 5, nf=NF_PCT, note='Thu từ năm thứ 2')

r += 1
section(ws, r, 'F. ĐẦU TƯ CAPEX (link từ sheet chi tiết)'); r += 1
gset('CAP-01', 'Nhân sự phát triển lõi', 'tỷ', None, note='← Nhan_Su (đội 10 người)')
gset('CAP-02', 'Trang thiết bị văn phòng', 'tỷ', None, note='← Thiet_Bi (Cloud-First)')
gset('CAP-03', 'Bản quyền & API tích hợp', 'tỷ', None, note='← CP_Khac (nhóm A)')
gset('CAP-04', 'Marketing & Sales ra mắt', 'tỷ', None, note='← CP_Khac (nhóm B)')
gset('CAP-05', 'Tư vấn, PM & Pháp lý', 'tỷ', None, note='← CP_Khac (nhóm C)')

r += 1
section(ws, r, 'G. CHI PHÍ VẬN HÀNH OPEX'); r += 1
gset('OPX-01', 'Nhân sự vận hành', 'tỷ', None, note='← Nhan_Su (đội vận hành)')
gset('OPX-02', 'Thuê hạ tầng Cloud trong nước', 'tỷ', None, note='← Cloud')
gset('OPX-03', 'Chi phí vận hành khác', 'tỷ', [0, 11.0, 22.0, 26.0, 31.0], survey=True,
     note='⚠️ Văn phòng, marketing thường niên, license API, pentest an ninh')

# =========================================================================
# 3. NHAN_SU
# =========================================================================
ns = wb.create_sheet('Nhan_Su')
ws_setup(ns, 'NHÂN SỰ & TIỀN LƯƠNG (BOTTOM-UP)', 'Đội phát triển (CAPEX) + đội vận hành (OPEX). Sửa ô vàng.')
ns.column_dimensions['B'].width = 42
ns.column_dimensions['C'].width = 13
rr = 4
section(ns, rr, 'A. THAM SỐ PHỤ PHÍ (đội phát triển)'); rr += 1
nsp = {}
for code, label, val in [('NS-L01', 'BHXH/BHYT/BHTN (% lương)', 0.215), ('NS-L02', 'Văn phòng /người/tháng (tr)', 10),
                         ('NS-L03', 'Thiết bị & công cụ /người/năm (tr)', 8), ('NS-L04', 'Phúc lợi & đào tạo /người/tháng (tr)', 2),
                         ('NS-L05', 'AI tools /người/tháng (tr)', 1.5)]:
    ns[f'A{rr}'] = code; ns[f'A{rr}'].font = f_bold; ns[f'A{rr}'].alignment = ctr
    ns[f'B{rr}'] = label; ns[f'B{rr}'].font = f_calc; ns[f'B{rr}'].alignment = left
    ns[f'D{rr}'] = val; ns[f'D{rr}'].font = f_input; ns[f'D{rr}'].fill = fill_input
    ns[f'D{rr}'].number_format = (NF_PCT if 'BHXH' in label else NF_TR); ns[f'D{rr}'].border = border; ns[f'D{rr}'].alignment = ctr
    nsp[code] = rr; rr += 1
P01 = f'$D${nsp["NS-L01"]}'; P02 = f'$D${nsp["NS-L02"]}'; P03 = f'$D${nsp["NS-L03"]}'
P04 = f'$D${nsp["NS-L04"]}'; P05 = f'$D${nsp["NS-L05"]}'


def team_table(title, roles, months_vals, total_label, mode, mult_ref=None):
    global rr
    rr += 1
    section(ns, rr, title); rr += 1
    ns[f'A{rr}'] = 'Mã'; ns[f'B{rr}'] = 'Vị trí'; ns[f'C{rr}'] = 'Lương/th (tr)'
    for i, c in enumerate(YCOL):
        ns[f'{c}{rr}'] = YEARS[i]
    for col in ['A', 'B', 'C'] + YCOL:
        ns[f'{col}{rr}'].font = f_hdr; ns[f'{col}{rr}'].fill = fill_hdr; ns[f'{col}{rr}'].alignment = ctr; ns[f'{col}{rr}'].border = border
    rr += 1
    rs = rr
    for code, label, sal, heads in roles:
        ns[f'A{rr}'] = code; ns[f'A{rr}'].font = f_bold; ns[f'A{rr}'].alignment = ctr; ns[f'A{rr}'].border = border
        ns[f'B{rr}'] = label; ns[f'B{rr}'].font = f_calc; ns[f'B{rr}'].alignment = left; ns[f'B{rr}'].border = border
        ns[f'C{rr}'] = sal; ns[f'C{rr}'].font = f_input; ns[f'C{rr}'].fill = fill_input; ns[f'C{rr}'].number_format = NF_TR; ns[f'C{rr}'].border = border; ns[f'C{rr}'].alignment = ctr
        for i, c in enumerate(YCOL):
            ns[f'{c}{rr}'] = heads[i]; cc = ns[f'{c}{rr}']; cc.font = f_input; cc.fill = fill_input; cc.number_format = NF_INT; cc.border = border; cc.alignment = ctr
        rr += 1
    re_ = rr - 1
    ns[f'B{rr}'] = 'Số tháng tính lương'; ns[f'B{rr}'].font = f_bold; ns[f'B{rr}'].alignment = left
    for i, c in enumerate(YCOL):
        ns[f'{c}{rr}'] = months_vals[i]; cc = ns[f'{c}{rr}']; cc.font = f_input; cc.fill = fill_input; cc.number_format = NF_INT; cc.border = border; cc.alignment = ctr
    rmonths = rr; rr += 1

    def agg(label, tpl, nf=NF_TY, bold=False, note=''):
        global rr
        ns[f'B{rr}'] = label; ns[f'B{rr}'].font = (f_bold if bold else f_calc); ns[f'B{rr}'].alignment = left
        for c in YCOL:
            ns[f'{c}{rr}'] = tpl.format(c=c); cc = ns[f'{c}{rr}']; cc.font = (Font(name=F, size=10, bold=True) if bold else f_calc)
            cc.number_format = nf; cc.border = border; cc.alignment = rgt
            if bold:
                cc.fill = fill_tot
        if note:
            ns[f'I{rr}'] = note; ns[f'I{rr}'].font = f_sub
        o = rr; rr += 1; return o

    r_heads = agg('Tổng nhân sự (người)', f'=SUM({{c}}{rs}:{{c}}{re_})', nf=NF_INT, bold=True)
    r_gross = agg('Lương gộp/năm', f'=SUMPRODUCT($C${rs}:$C${re_},{{c}}{rs}:{{c}}{re_})*{{c}}{rmonths}/1000', note='Σ(SL×lương)×tháng/1000')
    if mode == 'build':
        r_bhxh = agg('+ BHXH/BHYT/BHTN', f'={{c}}{r_gross}*{P01}')
        r_off = agg('+ Văn phòng', f'={{c}}{r_heads}*{P02}*{{c}}{rmonths}/1000')
        r_eq = agg('+ Thiết bị & công cụ', f'={{c}}{r_heads}*{P03}*{{c}}{rmonths}/12/1000')
        r_wf = agg('+ Phúc lợi & đào tạo', f'={{c}}{r_heads}*{P04}*{{c}}{rmonths}/1000')
        r_ai = agg('+ AI tools', f'={{c}}{r_heads}*{P05}*{{c}}{rmonths}/1000')
        comp = f'={{c}}{r_gross}+{{c}}{r_bhxh}+{{c}}{r_off}+{{c}}{r_eq}+{{c}}{r_wf}+{{c}}{r_ai}'
    else:  # ops: gross x he so tai
        comp = f'={{c}}{r_gross}*{{c}}{mult_ref}'
    r_all = agg(total_label, comp, bold=True, note=('all-in' if mode == 'build' else 'lương gộp × hệ số tải'))
    return r_all


build_roles = [
    ('NS-01', 'CTO / Kiến trúc trưởng (gộp Eng Director + Tech Lead)', 78, [1, 1, 0, 0, 0]),
    ('NS-02', 'Senior Kỹ sư BIM/IFC + WebGL (lõi đồ họa)', 55, [3, 3, 0, 0, 0]),
    ('NS-03', 'AI/ML + Data Engineer', 50, [1, 1, 0, 0, 0]),
    ('NS-04', 'Full-stack Backend / DevOps', 42, [2, 2, 0, 0, 0]),
    ('NS-05', 'Frontend / UX-UI Engineer', 36, [1, 1, 0, 0, 0]),
    ('NS-06', 'Security Engineer (QCVN 12)', 58, [1, 1, 0, 0, 0]),
    ('NS-07', 'Product Manager / BIM BA (gộp PM + BA + QA)', 55, [1, 1, 0, 0, 0]),
]
build_allin = team_table('B. ĐỘI NGŨ PHÁT TRIỂN (10 người, build 9/2026–giữa 2027 → CAP-01)',
                         build_roles, [4, 6, 0, 0, 0], '➜ NHÂN SỰ CAPEX (all-in) → CAP-01', 'build')

# ops: overhead multiplier link tu Gia_Dinh FIN-07
ops_roles = [
    ('NS-O1', 'Quản lý vận hành', 50, [0, 1, 1, 1, 1]),
    ('NS-O2', 'Bảo trì BIM / IFC Support', 40, [0, 1, 2, 2, 2]),
    ('NS-O3', 'Kỹ sư Backend / DevOps', 42, [0, 1, 2, 2, 2]),
    ('NS-O4', 'Kỹ sư Frontend', 36, [0, 1, 1, 1, 1]),
    ('NS-O5', 'Chuyên viên An ninh (Security Ops)', 45, [0, 1, 1, 1, 1]),
    ('NS-O6', 'Kỹ sư QA / Kiểm thử', 30, [0, 1, 1, 1, 1]),
    ('NS-O7', 'Nhân viên Sales (B2B/B2G)', 25, [0, 1, 3, 4, 4]),
    ('NS-O8', 'Chăm sóc KH (Customer Success)', 25, [0, 1, 3, 4, 4]),
    ('NS-O9', 'Hỗ trợ kỹ thuật (L1/L2 Support)', 20, [0, 2, 4, 4, 4]),
    ('NS-O10', 'Nhân sự G&A (HR, Kế toán, Admin)', 22, [0, 2, 2, 2, 2]),
]
# need multiplier per-col reference -> Gia_Dinh FIN-07 columns
mult_ref = f'{{c}}'  # placeholder
# We embed Gia_Dinh ref directly: replace in agg by building tpl with Gia_Dinh
# Simpler: pass a row ref that lives on this sheet. We'll add a hidden multiplier row referencing Gia_Dinh.
rr += 1
section(ns, rr, 'C. ĐỘI NGŨ VẬN HÀNH (12→20→22 người, từ giữa 2027 → OPX-01)'); rr += 1
# header
ns[f'A{rr}'] = 'Mã'; ns[f'B{rr}'] = 'Vị trí'; ns[f'C{rr}'] = 'Lương/th (tr)'
for i, c in enumerate(YCOL):
    ns[f'{c}{rr}'] = YEARS[i]
for col in ['A', 'B', 'C'] + YCOL:
    ns[f'{col}{rr}'].font = f_hdr; ns[f'{col}{rr}'].fill = fill_hdr; ns[f'{col}{rr}'].alignment = ctr; ns[f'{col}{rr}'].border = border
rr += 1
ors = rr
for code, label, sal, heads in ops_roles:
    ns[f'A{rr}'] = code; ns[f'A{rr}'].font = f_bold; ns[f'A{rr}'].alignment = ctr; ns[f'A{rr}'].border = border
    ns[f'B{rr}'] = label; ns[f'B{rr}'].font = f_calc; ns[f'B{rr}'].alignment = left; ns[f'B{rr}'].border = border
    ns[f'C{rr}'] = sal; ns[f'C{rr}'].font = f_input; ns[f'C{rr}'].fill = fill_input; ns[f'C{rr}'].number_format = NF_TR; ns[f'C{rr}'].border = border; ns[f'C{rr}'].alignment = ctr
    for i, c in enumerate(YCOL):
        ns[f'{c}{rr}'] = heads[i]; cc = ns[f'{c}{rr}']; cc.font = f_input; cc.fill = fill_input; cc.number_format = NF_INT; cc.border = border; cc.alignment = ctr
    rr += 1
ore = rr - 1
ns[f'B{rr}'] = 'Số tháng tính lương'; ns[f'B{rr}'].font = f_bold; ns[f'B{rr}'].alignment = left
for i, c in enumerate(YCOL):
    ns[f'{c}{rr}'] = [0, 6, 12, 12, 12][i]; cc = ns[f'{c}{rr}']; cc.font = f_input; cc.fill = fill_input; cc.number_format = NF_INT; cc.border = border; cc.alignment = ctr
omonths = rr; rr += 1
ns[f'B{rr}'] = 'Hệ số tải (← Gia_Dinh FIN-07)'; ns[f'B{rr}'].font = f_calc; ns[f'B{rr}'].alignment = left
for c in YCOL:
    ns[f'{c}{rr}'] = f'={G("FIN-07", c)}'; cc = ns[f'{c}{rr}']; cc.font = f_link; cc.number_format = NF_X; cc.border = border; cc.alignment = rgt
omult = rr; rr += 1
ns[f'B{rr}'] = 'Tổng nhân sự (người)'; ns[f'B{rr}'].font = f_bold; ns[f'B{rr}'].alignment = left
for c in YCOL:
    ns[f'{c}{rr}'] = f'=SUM({c}{ors}:{c}{ore})'; cc = ns[f'{c}{rr}']; cc.font = Font(name=F, size=10, bold=True); cc.number_format = NF_INT; cc.border = border; cc.alignment = rgt
oheads = rr; rr += 1
ns[f'B{rr}'] = 'Quỹ lương gộp/tháng (tr)'; ns[f'B{rr}'].font = f_calc; ns[f'B{rr}'].alignment = left
for c in YCOL:
    ns[f'{c}{rr}'] = f'=SUMPRODUCT($C${ors}:$C${ore},{c}{ors}:{c}{ore})'; cc = ns[f'{c}{rr}']; cc.font = f_calc; cc.number_format = NF_INT; cc.border = border; cc.alignment = rgt
ogross = rr; rr += 1
ns[f'B{rr}'] = '➜ NHÂN SỰ VẬN HÀNH (OPEX) → OPX-01'; ns[f'B{rr}'].font = f_bold; ns[f'B{rr}'].alignment = left
for c in YCOL:
    ns[f'{c}{rr}'] = f'={c}{ogross}*{c}{omonths}*{c}{omult}/1000'; cc = ns[f'{c}{rr}']; cc.font = Font(name=F, size=10, bold=True); cc.fill = fill_tot; cc.number_format = NF_TY; cc.border = border; cc.alignment = rgt
ns[f'I{rr}'] = 'quỹ lương × tháng × hệ số tải'; ns[f'I{rr}'].font = f_sub
ops_allin = rr

# link CAP-01 & OPX-01
gd = wb['Gia_Dinh']
for c in YCOL:
    gd[f'{c}{gi["CAP-01"]}'] = f'=Nhan_Su!{c}{build_allin}'; gd[f'{c}{gi["CAP-01"]}'].font = f_link; gd[f'{c}{gi["CAP-01"]}'].number_format = NF_TY; gd[f'{c}{gi["CAP-01"]}'].border = border; gd[f'{c}{gi["CAP-01"]}'].alignment = rgt
    gd[f'{c}{gi["OPX-01"]}'] = f'=Nhan_Su!{c}{ops_allin}'; gd[f'{c}{gi["OPX-01"]}'].font = f_link; gd[f'{c}{gi["OPX-01"]}'].number_format = NF_TY; gd[f'{c}{gi["OPX-01"]}'].border = border; gd[f'{c}{gi["OPX-01"]}'].alignment = rgt

# =========================================================================
# 4. THIET_BI (CAP-02, office equipment - Cloud-First)
# =========================================================================
tb = wb.create_sheet('Thiet_Bi')
ws_setup(tb, 'TRANG THIẾT BỊ VĂN PHÒNG (CAP-02)', 'Cloud-First: KHÔNG mua server (thuê cloud → OPEX). Chỉ thiết bị làm việc.')
tb.column_dimensions['B'].width = 42; tb.column_dimensions['C'].width = 13
rr = 4
section(tb, rr, 'THIẾT BỊ ĐỘI NGŨ (mua 1 lần)'); rr += 1
tb[f'A{rr}'] = 'Mã'; tb[f'B{rr}'] = 'Hạng mục'; tb[f'C{rr}'] = 'Đơn giá (tr)'
for i, c in enumerate(YCOL):
    tb[f'{c}{rr}'] = 'SL ' + YEARS[i]
tb[f'I{rr}'] = 'Ghi chú'
for col in ['A', 'B', 'C'] + YCOL + ['I']:
    tb[f'{col}{rr}'].font = f_hdr; tb[f'{col}{rr}'].fill = fill_hdr; tb[f'{col}{rr}'].alignment = ctr; tb[f'{col}{rr}'].border = border
rr += 1
tbs = rr
for code, label, unit, qty, note in [
    ('TB-01', 'Workstation chuyên dụng BIM/WebGL', 50, [3, 0, 0, 0, 0], 'Đội đồ họa 3D'),
    ('TB-02', 'Laptop cấu hình cao cho Dev', 35, [7, 1, 0, 0, 0], 'Đội phát triển'),
    ('TB-03', 'Thiết bị mạng & tường lửa văn phòng', 200, [1, 0, 0, 0, 0], 'An ninh local'),
    ('TB-04', 'Setup hạ tầng văn phòng & bổ sung', 105, [1, 3, 0, 0, 0], 'Bàn ghế, màn hình, mở rộng 2027'),
]:
    tb[f'A{rr}'] = code; tb[f'A{rr}'].font = f_bold; tb[f'A{rr}'].alignment = ctr; tb[f'A{rr}'].border = border
    tb[f'B{rr}'] = label; tb[f'B{rr}'].font = f_calc; tb[f'B{rr}'].alignment = left; tb[f'B{rr}'].border = border
    tb[f'C{rr}'] = unit; tb[f'C{rr}'].font = f_input; tb[f'C{rr}'].fill = fill_input; tb[f'C{rr}'].number_format = NF_TR; tb[f'C{rr}'].border = border; tb[f'C{rr}'].alignment = ctr
    for i, c in enumerate(YCOL):
        tb[f'{c}{rr}'] = qty[i]; cc = tb[f'{c}{rr}']; cc.font = f_input; cc.fill = fill_input; cc.number_format = NF_INT; cc.border = border; cc.alignment = ctr
    tb[f'I{rr}'] = note; tb[f'I{rr}'].font = f_sub
    rr += 1
tbe = rr - 1
tb[f'B{rr}'] = 'TỔNG THIẾT BỊ → CAP-02'; tb[f'B{rr}'].font = Font(name=F, size=11, bold=True); tb[f'B{rr}'].alignment = left
for c in YCOL:
    tb[f'{c}{rr}'] = f'=SUMPRODUCT($C${tbs}:$C${tbe},{c}{tbs}:{c}{tbe})/1000'; cc = tb[f'{c}{rr}']; cc.font = Font(name=F, size=11, bold=True); cc.fill = fill_kpi; cc.number_format = NF_TY; cc.border = border; cc.alignment = rgt
tb_total = rr
for c in YCOL:
    gd[f'{c}{gi["CAP-02"]}'] = f'=Thiet_Bi!{c}{tb_total}'; gd[f'{c}{gi["CAP-02"]}'].font = f_link; gd[f'{c}{gi["CAP-02"]}'].number_format = NF_TY; gd[f'{c}{gi["CAP-02"]}'].border = border; gd[f'{c}{gi["CAP-02"]}'].alignment = rgt

# =========================================================================
# 5. CP_KHAC (CAP-03/04/05)
# =========================================================================
ck = wb.create_sheet('CP_Khac')
ws_setup(ck, 'BẢN QUYỀN, MARKETING & TƯ VẤN (BOTTOM-UP)', 'Tổng từng nhóm → CAP-03/04/05. (đơn giá: triệu VNĐ)')
ck.column_dimensions['B'].width = 44; ck.column_dimensions['C'].width = 13
rr = 4


def ck_group(title, items):
    global rr
    section(ck, rr, title); rr += 1
    ck[f'A{rr}'] = 'Mã'; ck[f'B{rr}'] = 'Hạng mục'; ck[f'C{rr}'] = 'Đơn giá (tr)'
    for i, c in enumerate(YCOL):
        ck[f'{c}{rr}'] = 'SL ' + YEARS[i]
    ck[f'I{rr}'] = 'Ghi chú'
    for col in ['A', 'B', 'C'] + YCOL + ['I']:
        ck[f'{col}{rr}'].font = f_hdr; ck[f'{col}{rr}'].fill = fill_hdr; ck[f'{col}{rr}'].alignment = ctr; ck[f'{col}{rr}'].border = border
    rr += 1
    rs = rr
    for code, label, unit, qty, note in items:
        ck[f'A{rr}'] = code; ck[f'A{rr}'].font = f_bold; ck[f'A{rr}'].alignment = ctr; ck[f'A{rr}'].border = border
        ck[f'B{rr}'] = label; ck[f'B{rr}'].font = f_calc; ck[f'B{rr}'].alignment = left; ck[f'B{rr}'].border = border
        ck[f'C{rr}'] = unit; ck[f'C{rr}'].font = f_input; ck[f'C{rr}'].fill = fill_input; ck[f'C{rr}'].number_format = NF_TR; ck[f'C{rr}'].border = border; ck[f'C{rr}'].alignment = ctr
        for i, c in enumerate(YCOL):
            ck[f'{c}{rr}'] = qty[i]; cc = ck[f'{c}{rr}']; cc.font = f_input; cc.fill = fill_input; cc.number_format = NF_TR; cc.border = border; cc.alignment = ctr
        if note:
            ck[f'I{rr}'] = note; ck[f'I{rr}'].font = f_sub
        rr += 1
    re_ = rr - 1
    ck[f'B{rr}'] = 'TỔNG NHÓM'; ck[f'B{rr}'].font = Font(name=F, size=10, bold=True); ck[f'B{rr}'].alignment = left
    for c in YCOL:
        ck[f'{c}{rr}'] = f'=SUMPRODUCT($C${rs}:$C${re_},{c}{rs}:{c}{re_})/1000'; cc = ck[f'{c}{rr}']; cc.font = Font(name=F, size=10, bold=True); cc.fill = fill_kpi; cc.number_format = NF_TY; cc.border = border; cc.alignment = rgt
    sub = rr; rr += 2
    return sub


bq = ck_group('A. BẢN QUYỀN & API TÍCH HỢP (→ CAP-03)', [
    ('BQ-01', 'Bản quyền SDK BIM/IFC thương mại (/năm)', 2000, [1, 1, 0, 0, 0], 'ODA SDK, xử lý IFC'),
    ('BQ-02', 'Engine WebGL tối ưu hiển thị 3D (/năm)', 1000, [1, 1, 0, 0, 0], 'Render 3D nặng'),
])
mk = ck_group('B. MARKETING & SALES RA MẮT (→ CAP-04)', [
    ('MK-01', 'Nghiên cứu thị trường xây dựng 2026', 900, [1, 0, 0, 0, 0], 'Khảo sát thị trường'),
    ('MK-02', 'Hội thảo PMU/Sở XD (/sự kiện)', 500, [0, 6, 0, 0, 0], 'Giới thiệu giải pháp 2027'),
    ('MK-03', 'Chiến dịch marketing B2B trực tiếp', 1500, [0, 1, 0, 0, 0], 'Ra mắt 2027'),
])
pm = ck_group('C. TƯ VẤN, PM & PHÁP LÝ (→ CAP-05)', [
    ('PM-01', 'Tư vấn kỹ thuật & kiểm định QCVN 12', 1000, [1, 3, 0, 0, 0], 'Cao điểm kiểm định an ninh 2027'),
    ('PM-02', 'Đăng ký IP, pháp lý sản phẩm', 800, [1, 0, 0, 0, 0], 'Sở hữu trí tuệ (2026)'),
])
for c in YCOL:
    for code, sub in (('CAP-03', bq), ('CAP-04', mk), ('CAP-05', pm)):
        gd[f'{c}{gi[code]}'] = f'=CP_Khac!{c}{sub}'; gd[f'{c}{gi[code]}'].font = f_link; gd[f'{c}{gi[code]}'].number_format = NF_TY; gd[f'{c}{gi[code]}'].border = border; gd[f'{c}{gi[code]}'].alignment = rgt

# =========================================================================
# 6. CLOUD (OPX-02)
# =========================================================================
cl = wb.create_sheet('Cloud')
ws_setup(cl, 'CHI PHÍ THUÊ CLOUD TRONG NƯỚC (OPX-02)', 'Viettel/VNPT/CMC Cloud. Tăng theo số user & dung lượng BIM.')
cl.column_dimensions['B'].width = 42
hdr_row(cl, 3, 'Khảo sát / Giải trình')
rr = 4
section(cl, rr, 'HẠNG MỤC THUÊ CLOUD'); rr += 1
cls = rr
for code, label, vals, note in [
    ('CL-01', 'Server lưu trữ BIM & Data', [0.05, 0.40, 1.00, 1.80, 2.50], 'Object Storage ~900đ/GB/th, tích lũy tăng'),
    ('CL-02', 'Server xử lý WebGL & Node', [0.05, 0.40, 0.80, 1.20, 1.60], 'VM 16vCPU/64GB render server-side'),
    ('CL-03', 'Dịch vụ DevOps & Monitor', [0.05, 0.20, 0.30, 0.40, 0.60], 'Load Balancer, Firewall, CI/CD'),
    ('CL-04', 'Bản quyền OS, DB & SSL', [0.05, 0.20, 0.40, 0.60, 0.80], 'PostgreSQL Enterprise, SSL'),
]:
    setrow(cl, rr, code, label, 'tỷ', vals, kind='input', nf=NF_TR, note=note)
    rr += 1
cle = rr - 1
setrow(cl, rr, '', 'TỔNG OPEX THUÊ CLOUD → OPX-02', 'tỷ', [f'=SUM({c}{cls}:{c}{cle})' for c in YCOL], kind='calc', total=True)
cl_total = rr
for c in YCOL:
    gd[f'{c}{gi["OPX-02"]}'] = f'=Cloud!{c}{cl_total}'; gd[f'{c}{gi["OPX-02"]}'].font = f_link; gd[f'{c}{gi["OPX-02"]}'].number_format = NF_TY; gd[f'{c}{gi["OPX-02"]}'].border = border; gd[f'{c}{gi["OPX-02"]}'].alignment = rgt

# =========================================================================
# 7. DOANH_THU (SaaS + 3 On-Premise segments)
# =========================================================================
dt = {}
ws = wb.create_sheet('Doanh_Thu')
ws_setup(ws, 'DỰ PHÓNG DOANH THU 5 NĂM', 'SaaS + On-Premise (PMU, Sở XD, Doanh nghiệp). Đơn vị: tỷ VNĐ.')
hdr_row(ws, 3, 'Diễn giải')
r = 4


def dset(key, code, label, unit, vals, **kw):
    global r
    setrow(ws, r, code, label, unit, vals, **kw)
    dt[key] = r; r += 1


def onprem_block(seg_name, n_code, p_code, n_key, p_key):
    """Tao block On-Premise: contracts, price, license, cum, maint, total. Tra ve total row."""
    global r
    dset(n_key, '', f'{seg_name} — hợp đồng mới', 'HĐ', [f'={G(n_code, c)}' for c in YCOL], kind='link', nf=NF_INT)
    dset(p_key, '', f'{seg_name} — giá HĐ', 'tỷ', [f'={G(p_code, c)}' for c in YCOL], kind='link', nf=NF_TY)
    lic = r; dset('lic_' + n_key, '', f'{seg_name} — license mới', 'tỷ', [f'={c}{dt[n_key]}*{c}{dt[p_key]}' for c in YCOL], kind='calc')
    clic = r; dset('clic_' + n_key, '', f'{seg_name} — license lũy kế', 'tỷ', None, kind='calc')
    maint = r; dset('mn_' + n_key, '', f'{seg_name} — bảo trì (AMC)', 'tỷ', None, kind='calc')
    tot = r; dset('tot_' + n_key, '', f'{seg_name} — TỔNG', 'tỷ', [f'={c}{lic}+{c}{maint}' for c in YCOL], kind='calc', total=True)
    for i, c in enumerate(YCOL):
        prev = YCOL[i - 1] if i > 0 else None
        ws[f'{c}{clic}'] = (f'={c}{lic}' if i == 0 else f'={prev}{clic}+{c}{lic}')
        ws[f'{c}{maint}'] = (0 if i == 0 else f'={prev}{clic}*{G("AMC-01", c)}')
        for rowx, nf in ((clic, NF_TY), (maint, NF_TY)):
            cc = ws[f'{c}{rowx}']; cc.font = f_calc; cc.number_format = nf; cc.border = border; cc.alignment = rgt
    return tot


section(ws, r, 'KÊNH 1 — SaaS'); r += 1
dset('su', '', 'User active cuối kỳ', 'user', [f'={G("SAAS-01", c)}' for c in YCOL], kind='link', nf=NF_INT)
dset('sa', '', 'ARPU (tr/user/tháng)', 'tr', [f'={G("SAAS-02", c)}' for c in YCOL], kind='link', nf=NF_TR)
dset('sm', '', 'Số tháng thương mại', 'tháng', [f'={G("SAAS-03", c)}' for c in YCOL], kind='link', nf=NF_INT)
dset('srev', '', 'Doanh thu SaaS', 'tỷ', [f'={c}{dt["su"]}*{c}{dt["sa"]}*{c}{dt["sm"]}/1000' for c in YCOL], kind='calc', total=True, note='user×ARPU×tháng/1000')
r += 1
section(ws, r, 'KÊNH 2 — ON-PREMISE PMU'); r += 1
pmu_tot = onprem_block('PMU', 'OPP-01', 'OPP-02', 'pn', 'pp')
r += 1
section(ws, r, 'KÊNH 3 — ON-PREMISE SỞ XÂY DỰNG'); r += 1
sxd_tot = onprem_block('Sở XD', 'OSX-01', 'OSX-02', 'sn', 'sp')
r += 1
section(ws, r, 'KÊNH 4 — ON-PREMISE DOANH NGHIỆP'); r += 1
dn_tot = onprem_block('Doanh nghiệp', 'OPV-01', 'OPV-02', 'vn', 'vp')
r += 1
section(ws, r, 'TỔNG HỢP'); r += 1
dset('total', '', 'TỔNG DOANH THU', 'tỷ',
     [f'={c}{dt["srev"]}+{c}{pmu_tot}+{c}{sxd_tot}+{c}{dn_tot}' for c in YCOL], kind='calc', total=True, nf=NF_TY)
ws[f'B{dt["total"]}'].font = Font(name=F, size=11, bold=True)
DT_TOTAL = dt['total']

# =========================================================================
# 8. DAU_TU (CAPEX summary)
# =========================================================================
du = {}
ws = wb.create_sheet('Dau_Tu')
ws_setup(ws, 'TỔNG HỢP VỐN ĐẦU TƯ — CAPEX', 'Cloud-First, CIC 70% + NATIF 30%. Đơn vị: tỷ VNĐ.')
hdr_row(ws, 3, 'Nguồn')
r = 4
section(ws, r, 'HẠNG MỤC CAPEX (link Gia_Dinh)'); r += 1


def uset(key, code, label):
    global r
    setrow(ws, r, code, label, 'tỷ', [f'={G(code, c)}' for c in YCOL], kind='link')
    du[key] = r; r += 1


uset('c1', 'CAP-01', 'Nhân sự phát triển lõi')
uset('c2', 'CAP-02', 'Trang thiết bị văn phòng')
uset('c3', 'CAP-03', 'Bản quyền & API tích hợp')
uset('c4', 'CAP-04', 'Marketing & Sales ra mắt')
uset('c5', 'CAP-05', 'Tư vấn, PM & Pháp lý')
setrow(ws, r, '', 'TỔNG CAPEX/năm', 'tỷ', [f'=SUM({c}{du["c1"]}:{c}{du["c5"]})' for c in YCOL], kind='calc', total=True)
du['tot'] = r; r += 1
setrow(ws, r, '', 'Vốn CIC (70%)', 'tỷ', [f'={c}{du["tot"]}*{G("FIN-05", c)}' for c in YCOL], kind='calc', total=True)
du['cic'] = r; r += 1
setrow(ws, r, '', 'Vốn NATIF (30%)', 'tỷ', [f'={c}{du["tot"]}*{G("FIN-06", c)}' for c in YCOL], kind='calc', total=True)
du['nat'] = r
for k in ('tot', 'cic', 'nat'):
    ws[f'I{du[k]}'] = f'=SUM(D{du[k]}:H{du[k]})'; ws[f'I{du[k]}'].font = f_boldlink; ws[f'I{du[k]}'].number_format = NF_TY; ws[f'I{du[k]}'].alignment = rgt

# =========================================================================
# 9. OPEX (summary 3 lines)
# =========================================================================
op = {}
ws = wb.create_sheet('OPEX')
ws_setup(ws, 'TỔNG HỢP CHI PHÍ VẬN HÀNH — OPEX', 'Đơn vị: tỷ VNĐ.')
hdr_row(ws, 3, 'Nguồn')
r = 4
section(ws, r, 'HẠNG MỤC OPEX'); r += 1
setrow(ws, r, 'OPX-01', 'Nhân sự vận hành', 'tỷ', [f'={G("OPX-01", c)}' for c in YCOL], kind='link'); op['o1'] = r; r += 1
setrow(ws, r, 'OPX-02', 'Thuê hạ tầng Cloud', 'tỷ', [f'={G("OPX-02", c)}' for c in YCOL], kind='link'); op['o2'] = r; r += 1
setrow(ws, r, 'OPX-03', 'Chi phí vận hành khác', 'tỷ', [f'={G("OPX-03", c)}' for c in YCOL], kind='link'); op['o3'] = r; r += 1
setrow(ws, r, '', 'TỔNG OPEX/năm', 'tỷ', [f'=SUM({c}{op["o1"]}:{c}{op["o3"]})' for c in YCOL], kind='calc', total=True)
op['tot'] = r

# =========================================================================
# 10. P_AND_L
# =========================================================================
pl = {}
ws = wb.create_sheet('P_and_L')
ws_setup(ws, 'BÁO CÁO KẾT QUẢ & DÒNG TIỀN (P&L)', 'Đơn vị: tỷ VNĐ. Tất cả là công thức.')
hdr_row(ws, 3, 'Công thức')
r = 4


def pset(key, label, vals, note='', **kw):
    global r
    setrow(ws, r, '', label, 'tỷ', vals, note=note, **kw)
    pl[key] = r; r += 1


section(ws, r, 'KẾT QUẢ KINH DOANH'); r += 1
pset('rev', 'Doanh thu', [f'=Doanh_Thu!{c}{DT_TOTAL}' for c in YCOL], kind='link', note='← Doanh_Thu')
pset('gp', 'Lợi nhuận gộp (biên 60%)', [f'={c}{pl["rev"]}*{G("FIN-02", c)}' for c in YCOL], kind='calc', total=True)
pset('capex', 'Chi phí đầu tư (CAPEX)', [f'=-Dau_Tu!{c}{du["tot"]}' for c in YCOL], kind='link')
pset('opex', 'Chi phí vận hành (OPEX)', [f'=-OPEX!{c}{op["tot"]}' for c in YCOL], kind='link')
pset('net', 'DÒNG TIỀN RÒNG DỰ ÁN', [f'={c}{pl["gp"]}+{c}{pl["capex"]}+{c}{pl["opex"]}' for c in YCOL], kind='calc', total=True, note='LN gộp – CAPEX – OPEX')
pset('cum', 'Lũy kế dự án', None, kind='calc', total=True)
for i, c in enumerate(YCOL):
    prev = YCOL[i - 1] if i > 0 else None
    ws[f'{c}{pl["cum"]}'] = (f'={c}{pl["net"]}' if i == 0 else f'={prev}{pl["cum"]}+{c}{pl["net"]}')
    cc = ws[f'{c}{pl["cum"]}']; cc.font = Font(name=F, size=10, bold=True); cc.number_format = NF_TY; cc.border = border; cc.alignment = rgt; cc.fill = fill_tot

# =========================================================================
# 11. DONG_TIEN
# =========================================================================
cf = {}
ws = wb.create_sheet('Dong_Tien')
ws_setup(ws, 'DÒNG TIỀN DỰ ÁN & CIC', 'Đơn vị: tỷ VNĐ.')
hdr_row(ws, 3, 'Ghi chú')
r = 4


def cset(key, label, vals, note='', **kw):
    global r
    setrow(ws, r, '', label, 'tỷ', vals, note=note, **kw)
    cf[key] = r; r += 1


section(ws, r, 'GÓC ĐỘ TOÀN DỰ ÁN'); r += 1
cset('pnet', 'Dòng tiền ròng dự án', [f'=P_and_L!{c}{pl["net"]}' for c in YCOL], kind='link', total=True)
cset('pcum', 'Lũy kế dự án', [f'=P_and_L!{c}{pl["cum"]}' for c in YCOL], kind='link', total=True)
r += 1
section(ws, r, 'GÓC ĐỘ CIC (vốn đối ứng 70%)'); r += 1
cset('cnet', 'Dòng tiền ròng CIC', [f'=P_and_L!{c}{pl["gp"]}-Dau_Tu!{c}{du["cic"]}+P_and_L!{c}{pl["opex"]}' for c in YCOL], kind='calc', total=True, note='LN gộp – vốn CIC – OPEX')
cset('ccum', 'Lũy kế CIC', None, kind='calc', total=True)
for i, c in enumerate(YCOL):
    prev = YCOL[i - 1] if i > 0 else None
    ws[f'{c}{cf["ccum"]}'] = (f'={c}{cf["cnet"]}' if i == 0 else f'={prev}{cf["ccum"]}+{c}{cf["cnet"]}')
    cc = ws[f'{c}{cf["ccum"]}']; cc.font = Font(name=F, size=10, bold=True); cc.number_format = NF_TY; cc.border = border; cc.alignment = rgt; cc.fill = fill_tot

# =========================================================================
# 12. NPV_IRR (2026 = t0)
# =========================================================================
ws = wb.create_sheet('NPV_IRR')
ws_setup(ws, 'ĐÁNH GIÁ HIỆU QUẢ ĐẦU TƯ (NPV / IRR / PAYBACK)', 'Quy ước: 2026 = năm gốc (t=0). WACC ← Gia_Dinh.')
hdr_row(ws, 3, 'Giá trị')
r = 4
ni = {}
section(ws, r, 'DÒNG TIỀN'); r += 1
setrow(ws, r, '', 'Dòng tiền ròng dự án', 'tỷ', [f'=Dong_Tien!{c}{cf["pnet"]}' for c in YCOL], kind='link'); ni['pcf'] = r; r += 1
setrow(ws, r, '', 'Dòng tiền ròng CIC', 'tỷ', [f'=Dong_Tien!{c}{cf["cnet"]}' for c in YCOL], kind='link'); ni['ccf'] = r; r += 1
r += 1
section(ws, r, 'CHỈ SỐ HIỆU QUẢ (WACC, 2026=t0)'); r += 1
wacc = G('FIN-03', 'D')
metrics = [
    ('NPV toàn dự án (5 năm)', f'=D{ni["pcf"]}+NPV({wacc},E{ni["pcf"]}:H{ni["pcf"]})', NF_TY),
    ('IRR toàn dự án', f'=IFERROR(IRR(D{ni["pcf"]}:H{ni["pcf"]},0.2),"n/a")', NF_PCT),
    ('NPV đầu tư CIC (5 năm)', f'=D{ni["ccf"]}+NPV({wacc},E{ni["ccf"]}:H{ni["ccf"]})', NF_TY),
    ('IRR đầu tư CIC', f'=IFERROR(IRR(D{ni["ccf"]}:H{ni["ccf"]},0.2),"n/a")', NF_PCT),
]
for label, formula, nf in metrics:
    ws[f'B{r}'] = label; ws[f'B{r}'].font = f_bold; ws[f'B{r}'].alignment = left
    ws[f'D{r}'] = formula; ws[f'D{r}'].font = Font(name=F, size=11, bold=True); ws[f'D{r}'].number_format = nf
    ws[f'D{r}'].fill = fill_kpi; ws[f'D{r}'].border = border; ws[f'D{r}'].alignment = ctr
    r += 1
# payback (project & CIC), cumulative crosses 0
pcum = f'Dong_Tien!D{cf["pcum"]}:H{cf["pcum"]}'; pnet = f'Dong_Tien!D{cf["pnet"]}:H{cf["pnet"]}'
ccum = f'Dong_Tien!D{cf["ccum"]}:H{cf["ccum"]}'; cnet = f'Dong_Tien!D{cf["cnet"]}:H{cf["cnet"]}'


def payback(cum_ref, net_ref):
    n = f'COUNTIF({cum_ref},"<0")'
    return f'=IF({n}=0,0,IF({n}>=5,"Sau 5 năm",{n}-1+(-INDEX({cum_ref},{n}))/INDEX({net_ref},{n}+1)))'


for label, formula in [('Hoàn vốn dự án (năm thứ)', payback(pcum, pnet)), ('Hoàn vốn CIC (năm thứ)', payback(ccum, cnet))]:
    ws[f'B{r}'] = label; ws[f'B{r}'].font = f_bold; ws[f'B{r}'].alignment = left
    ws[f'D{r}'] = formula; ws[f'D{r}'].font = Font(name=F, size=11, bold=True); ws[f'D{r}'].number_format = '0.0'
    ws[f'D{r}'].fill = fill_kpi; ws[f'D{r}'].border = border; ws[f'D{r}'].alignment = ctr
    ws[f'I{r}'] = 'Số năm kể từ đầu 2026 (≈)'; ws[f'I{r}'].font = f_sub
    r += 1

# =========================================================================
# 13. HOA_VON
# =========================================================================
ws = wb.create_sheet('Hoa_Von')
ws_setup(ws, 'PHÂN TÍCH ĐIỂM HÒA VỐN VẬN HÀNH', 'Trạng thái vận hành đầy đủ (2030).')
ws.column_dimensions['B'].width = 44
r = 4
section(ws, r, 'HÒA VỐN THEO DOANH THU (2030)'); r += 1
hv = {}


def hrow(label, formula, nf, note=''):
    global r
    ws[f'B{r}'] = label; ws[f'B{r}'].font = f_calc; ws[f'B{r}'].alignment = left
    ws[f'D{r}'] = formula; ws[f'D{r}'].font = f_calc; ws[f'D{r}'].number_format = nf; ws[f'D{r}'].border = border; ws[f'D{r}'].alignment = ctr
    if note:
        ws[f'I{r}'] = note; ws[f'I{r}'].font = f_sub
    hv[label] = r; r += 1


hrow('Chi phí cố định/năm (OPEX 2030)', f'=OPEX!H{op["tot"]}', NF_TY, '← OPEX 2030')
hrow('Biên lợi nhuận gộp', f'={G("FIN-02","D")}', NF_PCT)
fc = hv['Chi phí cố định/năm (OPEX 2030)']; gm = hv['Biên lợi nhuận gộp']
hrow('Doanh thu hòa vốn/năm', f'=D{fc}/D{gm}', NF_TY, '= chi phí cố định / biên LN gộp')
hrow('Doanh thu thực tế 2030', f'=Doanh_Thu!H{DT_TOTAL}', NF_TY, 'Vượt xa ngưỡng hòa vốn')

# =========================================================================
# 14. DO_NHAY
# =========================================================================
ws = wb.create_sheet('Do_Nhay')
ws_setup(ws, 'PHÂN TÍCH ĐỘ NHẠY — 3 KỊCH BẢN', 'Sửa hệ số (ô vàng) để xem tác động NPV dự án.')
ws.column_dimensions['B'].width = 26
r = 4
section(ws, r, 'A. HỆ SỐ KỊCH BẢN (ô vàng)', ('A', 'D')); r += 1
ws[f'B{r}'] = 'Kịch bản'; ws[f'C{r}'] = 'Hệ số doanh thu'; ws[f'D{r}'] = 'Hệ số OPEX'
for col in 'BCD':
    ws[f'{col}{r}'].font = f_hdr; ws[f'{col}{r}'].fill = fill_hdr; ws[f'{col}{r}'].alignment = ctr; ws[f'{col}{r}'].border = border
r += 1
scen = {}
for name, rm, om in [('Bi quan', 0.7, 1.15), ('Cơ sở', 1.0, 1.0), ('Lạc quan', 1.25, 0.95)]:
    ws[f'B{r}'] = name; ws[f'B{r}'].font = f_bold; ws[f'B{r}'].border = border; ws[f'B{r}'].alignment = left
    ws[f'C{r}'] = rm; ws[f'D{r}'] = om
    for col in 'CD':
        ws[f'{col}{r}'].font = f_input; ws[f'{col}{r}'].fill = fill_input; ws[f'{col}{r}'].number_format = NF_X; ws[f'{col}{r}'].border = border; ws[f'{col}{r}'].alignment = ctr
    scen[name] = r; r += 1
r += 1
section(ws, r, 'B. NPV DỰ ÁN THEO KỊCH BẢN (2026=t0)'); r += 1
ws[f'B{r}'] = 'Kịch bản'
for i, c in enumerate(YCOL):
    ws[f'{c}{r}'] = YEARS[i]
ws[f'I{r}'] = 'NPV dự án'
for col in ['B'] + YCOL + ['I']:
    ws[f'{col}{r}'].font = f_hdr; ws[f'{col}{r}'].fill = fill_hdr; ws[f'{col}{r}'].alignment = ctr; ws[f'{col}{r}'].border = border
r += 1
gmref = G('FIN-02', 'D')
for name in ['Bi quan', 'Cơ sở', 'Lạc quan']:
    sr = scen[name]
    ws[f'B{r}'] = name; ws[f'B{r}'].font = f_bold; ws[f'B{r}'].border = border; ws[f'B{r}'].alignment = left
    for c in YCOL:
        ws[f'{c}{r}'] = (f'=Doanh_Thu!{c}{DT_TOTAL}*$C${sr}*{gmref}-Dau_Tu!{c}{du["tot"]}-OPEX!{c}{op["tot"]}*$D${sr}')
        cc = ws[f'{c}{r}']; cc.font = f_calc; cc.number_format = NF_TY; cc.border = border; cc.alignment = rgt
    ws[f'I{r}'] = f'=D{r}+NPV({wacc},E{r}:H{r})'
    ws[f'I{r}'].font = Font(name=F, size=10, bold=True); ws[f'I{r}'].number_format = NF_TY; ws[f'I{r}'].fill = fill_kpi; ws[f'I{r}'].border = border; ws[f'I{r}'].alignment = ctr
    r += 1

# =========================================================================
# 15. DASHBOARD
# =========================================================================
ws = wb.create_sheet('Dashboard')
ws_setup(ws, 'BẢNG TỔNG HỢP CHỈ SỐ (DASHBOARD)', 'Tự động liên kết từ các sheet.')
ws.column_dimensions['B'].width = 40
r = 4
section(ws, r, 'A. CHỈ SỐ DỰ ÁN'); r += 1
for label, formula, nf in [
    ('Tổng vốn đầu tư CAPEX (tỷ)', f'=Dau_Tu!I{du["tot"]}', NF_TY),
    ('Vốn CIC đối ứng 70% (tỷ)', f'=Dau_Tu!I{du["cic"]}', NF_TY),
    ('Tài trợ NATIF 30% (tỷ)', f'=Dau_Tu!I{du["nat"]}', NF_TY),
    ('NPV toàn dự án (tỷ)', f'=D{ni["pcf"]+1-1}', NF_TY),
]:
    pass
dash = [
    ('Tổng vốn đầu tư CAPEX (tỷ)', f'=Dau_Tu!I{du["tot"]}', NF_TY),
    ('Vốn CIC đối ứng 70% (tỷ)', f'=Dau_Tu!I{du["cic"]}', NF_TY),
    ('Tài trợ NATIF 30% (tỷ)', f'=Dau_Tu!I{du["nat"]}', NF_TY),
    ('NPV toàn dự án 5 năm (tỷ)', f'=NPV_IRR!D{ni["pcf"]}+NPV({wacc},NPV_IRR!E{ni["pcf"]}:H{ni["pcf"]})', NF_TY),
    ('IRR toàn dự án', f'=IFERROR(IRR(NPV_IRR!D{ni["pcf"]}:H{ni["pcf"]},0.2),"n/a")', NF_PCT),
    ('NPV đầu tư CIC (tỷ)', f'=NPV_IRR!D{ni["ccf"]}+NPV({wacc},NPV_IRR!E{ni["ccf"]}:H{ni["ccf"]})', NF_TY),
    ('IRR đầu tư CIC', f'=IFERROR(IRR(NPV_IRR!D{ni["ccf"]}:H{ni["ccf"]},0.2),"n/a")', NF_PCT),
]
for label, formula, nf in dash:
    ws[f'B{r}'] = label; ws[f'B{r}'].font = f_bold; ws[f'B{r}'].alignment = left; ws[f'B{r}'].border = border
    ws[f'D{r}'] = formula; ws[f'D{r}'].font = Font(name=F, size=11, bold=True, color=GREEN); ws[f'D{r}'].number_format = nf
    ws[f'D{r}'].fill = fill_kpi; ws[f'D{r}'].border = border; ws[f'D{r}'].alignment = ctr
    r += 1
r += 1
section(ws, r, 'B. DOANH THU & DÒNG TIỀN THEO NĂM (tỷ)'); r += 1
ws[f'B{r}'] = ''
for i, c in enumerate(YCOL):
    ws[f'{c}{r}'] = YEARS[i]; ws[f'{c}{r}'].font = f_hdr; ws[f'{c}{r}'].fill = fill_hdr; ws[f'{c}{r}'].alignment = ctr; ws[f'{c}{r}'].border = border
ws[f'B{r}'].fill = fill_hdr; ws[f'B{r}'].border = border
r += 1
for label, ref in [('Doanh thu', f'Doanh_Thu!{{c}}{DT_TOTAL}'), ('Dòng tiền ròng dự án', f'P_and_L!{{c}}{pl["net"]}'),
                   ('Lũy kế dự án', f'P_and_L!{{c}}{pl["cum"]}'), ('Lũy kế CIC', f'Dong_Tien!{{c}}{cf["ccum"]}')]:
    ws[f'B{r}'] = label; ws[f'B{r}'].font = f_bold; ws[f'B{r}'].alignment = left; ws[f'B{r}'].border = border
    for c in YCOL:
        ws[f'{c}{r}'] = '=' + ref.format(c=c); cc = ws[f'{c}{r}']; cc.font = f_link; cc.number_format = NF_TY; cc.border = border; cc.alignment = rgt
    r += 1

# =========================================================================
# 16. THUAT_NGU (glossary - compact)
# =========================================================================
tn = wb.create_sheet('Thuat_Ngu')
tn.sheet_view.showGridLines = False
tn['A1'] = 'CHÚ THÍCH THUẬT NGỮ & TỪ VIẾT TẮT'; tn['A1'].font = f_title
tn.column_dimensions['A'].width = 20; tn.column_dimensions['B'].width = 40; tn.column_dimensions['C'].width = 82
rr = 3
gloss = [
    ('A. TÀI CHÍNH', [('NPV', 'Net Present Value', 'Giá trị hiện tại thuần (>0 là sinh lời).'),
                      ('IRR', 'Internal Rate of Return', 'Tỷ suất sinh lợi nội bộ; so với WACC.'),
                      ('WACC', 'Weighted Average Cost of Capital', 'Chi phí vốn bình quân (chiết khấu 12%).'),
                      ('CAPEX', 'Capital Expenditure', 'Chi phí đầu tư ban đầu.'),
                      ('OPEX', 'Operating Expenditure', 'Chi phí vận hành thường xuyên.'),
                      ('COGS', 'Cost of Goods Sold', 'Giá vốn dịch vụ (40% doanh thu).'),
                      ('AMC', 'Annual Maintenance Contract', 'Phí bảo trì hàng năm (15% giá HĐ lũy kế).'),
                      ('ARPU', 'Average Revenue Per User', 'Doanh thu bình quân/user.'),
                      ('P&L', 'Profit and Loss', 'Báo cáo kết quả kinh doanh.')]),
    ('B. KINH DOANH', [('SaaS', 'Software as a Service', 'Phần mềm thuê bao trên cloud.'),
                       ('On-Premise', '', 'Cài đặt tại hạ tầng khách hàng.'),
                       ('B2G / B2B', 'Business to Government / Business', 'Bán cho nhà nước / doanh nghiệp.'),
                       ('PMU', 'Project Management Unit', 'Ban Quản lý dự án.'),
                       ('MVP', 'Minimum Viable Product', 'Sản phẩm khả dụng tối thiểu.')]),
    ('C. CÔNG NGHỆ', [('CDE', 'Common Data Environment', 'Môi trường dữ liệu chung.'),
                      ('BIM', 'Building Information Modeling', 'Mô hình thông tin công trình.'),
                      ('IFC', 'Industry Foundation Classes', 'Định dạng mở dữ liệu BIM (ISO 16739).'),
                      ('WebGL', 'Web Graphics Library', 'Đồ họa 3D trên trình duyệt.'),
                      ('API', 'Application Programming Interface', 'Giao diện kết nối hệ thống.'),
                      ('SDK', 'Software Development Kit', 'Bộ công cụ phát triển phần mềm.'),
                      ('CDN', 'Content Delivery Network', 'Mạng phân phối nội dung (tăng tốc tải).'),
                      ('CI/CD', 'Continuous Integration / Deployment', 'Tích hợp & triển khai liên tục.')]),
    ('D. AN NINH & PHÁP LÝ', [('QCVN 12', '', 'Quy chuẩn kỹ thuật quốc gia về an ninh mạng.'),
                              ('NATIF', 'National Technology Innovation Foundation', 'Quỹ Đổi mới công nghệ quốc gia.'),
                              ('NĐ 13/2023', '', 'Nghị định về bảo vệ dữ liệu cá nhân.'),
                              ('G&A', 'General & Administrative', 'Chi phí quản lý doanh nghiệp.'),
                              ('IP', 'Intellectual Property', 'Sở hữu trí tuệ.'),
                              ('BHXH', '', 'Bảo hiểm xã hội (21,5% lương).')]),
]
for sec_t, rows_ in gloss:
    tn[f'A{rr}'] = sec_t
    for col in 'ABC':
        tn[f'{col}{rr}'].fill = fill_sec
    tn[f'A{rr}'].font = f_sec
    rr += 1
    tn[f'A{rr}'] = 'Viết tắt'; tn[f'B{rr}'] = 'Tiếng Anh'; tn[f'C{rr}'] = 'Giải thích'
    for col in 'ABC':
        tn[f'{col}{rr}'].font = f_hdr; tn[f'{col}{rr}'].fill = fill_hdr; tn[f'{col}{rr}'].alignment = ctr; tn[f'{col}{rr}'].border = border
    rr += 1
    for term, eng, vn in rows_:
        tn[f'A{rr}'] = term; tn[f'A{rr}'].font = f_bold; tn[f'A{rr}'].border = border; tn[f'A{rr}'].alignment = Alignment(vertical='top')
        tn[f'B{rr}'] = eng; tn[f'B{rr}'].font = Font(name=F, size=9, italic=True, color='595959'); tn[f'B{rr}'].border = border; tn[f'B{rr}'].alignment = wrap
        tn[f'C{rr}'] = vn; tn[f'C{rr}'].font = f_calc; tn[f'C{rr}'].border = border; tn[f'C{rr}'].alignment = wrap
        rr += 1
    rr += 1

# Order & save
order = ['Huong_Dan', 'Thuat_Ngu', 'Gia_Dinh', 'Nhan_Su', 'Thiet_Bi', 'CP_Khac', 'Cloud', 'Doanh_Thu',
         'Dau_Tu', 'OPEX', 'P_and_L', 'Dong_Tien', 'NPV_IRR', 'Hoa_Von', 'Do_Nhay', 'Dashboard']
wb._sheets.sort(key=lambda s: order.index(s.title))
wb['Gia_Dinh'].freeze_panes = 'D4'
wb['Doanh_Thu'].freeze_panes = 'D4'
wb.calculation.fullCalcOnLoad = True
out = 'D:\\01_Projects\\cic-cde\\Ke hoach tai chinh CDE CIC (mo hinh dong).xlsx'
wb.save(out)
import sys
sys.stdout.buffer.write(('Saved: ' + out + '\n').encode('utf-8'))
